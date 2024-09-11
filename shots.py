import numpy as np
import os, sys
from datetime import datetime

sys.path.insert(1, './shared')

import streamlit as st
import pandas as pd

import base64
import boto3
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

# Use the Access Key and Secret Key you just created

AWS_ACCESS_KEY = st.secrets["aws"]["AWS_ACCESS_KEY"]
AWS_SECRET_KEY = st.secrets["aws"]["AWS_SECRET_KEY"]

S3_BUCKET = 'stevensonhockeydata'
EXCEL_FILE_KEY = 'Stevenson_Hockey.xlsx' 
#EXCEL_FILE_KEY = 'Stevenson_Hockey_Testing.xlsx' 


# Create an S3 client
s3 = boto3.client('s3', aws_access_key_id=AWS_ACCESS_KEY, aws_secret_access_key=AWS_SECRET_KEY)



def read_excel_from_s3(bucket, file_key, sheet_name):
    # Get the object from S3
    obj = s3.get_object(Bucket=bucket, Key=file_key)
    
    # Read the content of the file into a Pandas DataFrame
    excel_data = obj['Body'].read()
    df = pd.read_excel(BytesIO(excel_data), sheet_name=sheet_name)
    
    return df


roster_df = read_excel_from_s3(S3_BUCKET, EXCEL_FILE_KEY, sheet_name="Roster")



# Read the existing Excel file from S3
def read_excel_from_s3(bucket, file_key):
    obj = s3.get_object(Bucket=bucket, Key=file_key)
    return obj['Body'].read()

# Append new data to the "Shots" worksheet
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

#def append_to_excel_s3(bucket, file_key, data_to_save, sheet_name="Shots"):
def append_to_excel_s3(bucket, file_key, data_to_save, sheet_name):
    # Convert the incoming data (likely a list) to a DataFrame
    df_new = pd.DataFrame(data_to_save)

    # Load the existing Excel file from S3
    excel_data = read_excel_from_s3(bucket, file_key)
    
    # Open the workbook
    with BytesIO(excel_data) as buffer:
        workbook = load_workbook(buffer)
        writer_buffer = BytesIO()  # Buffer for saving the updated Excel file

        # Load existing data from "Shots" worksheet (if it exists)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Convert existing data in "Shots" to a DataFrame
            data = list(sheet.values)  # Convert the iterator to a list of rows
            if len(data) > 0:
                # The first row is treated as column names, the rest as data
                existing_df = pd.DataFrame(data[1:], columns=data[0])
            else:
                existing_df = pd.DataFrame()  # No data, initialize an empty DataFrame
        else:
            # Create a new sheet if it does not exist
            sheet = workbook.create_sheet(sheet_name)
            existing_df = pd.DataFrame()  # No existing data since the sheet is new

        # Ensure 'GameDate' is in both dataframes and convert to datetime
        if 'GameDate' in df_new.columns and 'GameDate' in existing_df.columns:
            # Convert 'GameDate' to datetime format in both DataFrames
            df_new['GameDate'] = pd.to_datetime(df_new['GameDate'], errors='coerce').dt.strftime('%Y-%m-%d')
            existing_df['GameDate'] = pd.to_datetime(existing_df['GameDate'], errors='coerce').dt.strftime('%Y-%m-%d')

        # Ensure no index column exists in existing_df
        if 'index' in existing_df.columns:
            existing_df = existing_df.drop(columns=['index'])

        # Ensure both DataFrames have the same columns and reset index
        if not existing_df.empty:
            df_new = df_new.reindex(columns=existing_df.columns)
            existing_df.reset_index(drop=True, inplace=True)

        # Check if df_new is a valid DataFrame and not empty
        if isinstance(df_new, pd.DataFrame) and not df_new.empty:
            # Concatenate the new data to the existing data
            updated_df = pd.concat([existing_df, df_new], ignore_index=True)
        else:
            updated_df = existing_df  # If no new data, keep existing data


            
        #st.dataframe(existing_df, width=650)
        #st.dataframe(df_new, width=650)
        #st.dataframe(updated_df, width=650)              
            
            
            
        # Clear the sheet and write the updated DataFrame back to the "Shots" sheet
        sheet.delete_rows(1, sheet.max_row)  # Clear the sheet

        # Write the updated DataFrame back to the sheet, including headers
        for r_idx, row in updated_df.iterrows():
            for c_idx, value in enumerate(row):
                sheet.cell(row=r_idx+2, column=c_idx+1, value=value)  # Write values starting from row 2
        for col_idx, col_name in enumerate(updated_df.columns):
            sheet.cell(row=1, column=col_idx+1, value=col_name)  # Write column headers

        # Save the updated workbook to the in-memory buffer
        workbook.save(writer_buffer)
        writer_buffer.seek(0)  # Go to the start of the buffer

        # Write the updated Excel file back to S3
        s3.put_object(Bucket=bucket, Key=file_key, Body=writer_buffer.getvalue())

        
        
        
# Placeholder for the data to be saved
data_to_save = []
faceoff_to_s3 = []
scores_to_s3 = []
goalie_to_s3 = []

#st.set_option('deprecation.showPyplotGlobalUse', False)


# Function to load an image and convert it to base64
def get_image_base64(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode()

# Image URL
logo_url = "./stevenson_logo.png"

# Get the base64 string of the image
logo_base64 = get_image_base64(logo_url)



# Page configuration
st.set_page_config(
    page_title="Stevenson Hockey Shots Data",
    page_icon=f"data:image/png;base64,{logo_base64}",
    layout="wide",
    initial_sidebar_state="expanded")



# HTML and CSS for centering the image
st.sidebar.markdown(
    f"""
    <div style="display: flex; justify-content: center;">
        <img src="data:image/png;base64,{logo_base64}" width="100">
    </div>
    """,
    unsafe_allow_html=True
)




# Load your data
#file_path = 'Stevenson_Hockey.xlsx'
#data = pd.ExcelFile('Stevenson_Hockey.xlsx')


# Load the sheets into separate DataFrames
#roster_df = pd.read_excel(data, sheet_name='Roster')

#######################


# Sidebar
with st.sidebar:
    st.subheader('Stevenson Hockey Game Data')
    
    # Select team from roster
    team_list = list(roster_df.Team.unique())[::-1]
    selected_team = st.selectbox('Select a team', team_list)
    
    # Filter and sort the selected team data
    df_selected_team = roster_df[roster_df.Team == selected_team]
    df_selected_team_sorted = df_selected_team.sort_values(by="JerseyNumber", ascending=False)
    unique_jersey_numbers = sorted(df_selected_team_sorted['JerseyNumber'].unique())
    unique_jersey_numbers = [None] + unique_jersey_numbers 

    # Add the option to add a new opponent
    opponents = ["Carmel","Fenwick","Glenbrook North","Glenbrook South","Lake Forest","Loyola","New Trier", "St. Ignatius", "St. Viator","York"]
    
    # If a new opponent is added, handle the input and update the opponent list
    opponent = st.selectbox("Select Opponent", opponents + ["Add New Opponent..."])

    if opponent == "Add New Opponent...":
        new_opponent = st.text_input("Enter new opponent name")
        if new_opponent:
            opponents.append(new_opponent)  # Dynamically add the new opponent to the list
            opponent = new_opponent  # Update the opponent variable to the new entry
            
            
    # Add game date selection with default to today's date
    game_date = st.date_input("Select Game Date", datetime.today())
            
        
    # Add a radio button to select if Stevenson is home or away
    stevenson_home = st.radio("Is Stevenson the Home Team?", ("Yes", "No"))





# Main area ###########################

# Main area
st.title("Hockey Game Scores Input")

with st.expander("Game Scores Input", expanded=False):

    st.subheader(f"Final Scores")
    cols = st.columns(2)  # Now, cols[0], cols[1], and cols[2] are valid

    with cols[0]:
        user_input = st.text_input("Stevenson Final Score", key=f"Score_Stevenson")

        # Initialize win_count as None
        Score_Stevenson = 0

        # Validate if the input is a number
        if user_input:
            if user_input.isdigit():
                Score_Stevenson = int(user_input)  # Convert the input to an integer and store in win_count
            else:
                st.error("Please enter only numbers.")
                
    with cols[1]:
        user_input = st.text_input(f"{opponent} Final Score", key=f"Score_Opponent")

        # Initialize win_count as None
        Score_Opponent = 0

        # Validate if the input is a number
        if user_input:
            if user_input.isdigit():
                Score_Opponent = int(user_input)  # Convert the input to an integer and store in win_count
            else:
                st.error("Please enter only numbers.")        
        
        
    if Score_Stevenson > Score_Opponent:
        Game_Win = "Yes"
    elif Score_Stevenson < Score_Opponent:
        Game_Win = "No"
    else:
        Game_Win = "Tie"        
        
   
        
        
    st.subheader(f"Player Scores")
    # Add a radio button for selecting the period, displayed horizontally
    period_game = st.radio("Select Period", options=["1", "2", "3", "Overtime"], horizontal=True, key="period_game")
    stevenson_game = st.slider(f"Scores by Stevenson in Period {period_game}", min_value=0, max_value=10, value=0)

    if stevenson_game > 0:
        st.subheader(f"Stevenson Scores Details for Period {period_game}")
        for i in range(stevenson_game):
            # Define 3 columns, with equal or customizable widths
            cols = st.columns(3)  # Now, cols[0], cols[1], and cols[2] are valid

            with cols[0]:
                score_jersey_number = st.selectbox(f"Jersey Number (score {i+1})", unique_jersey_numbers, key=f"score_jersey_{i}")

            with cols[1]:
                assistant_1_jersey = st.selectbox("Assistant_1", unique_jersey_numbers, key=f"Assistant_1_{i}")

            with cols[2]:
                assistant_2_jersey = st.selectbox("Assistant_2", unique_jersey_numbers, key=f"Assistant_2_{i}")

                
            if assistant_1_jersey == 0:
                assistant_1_jersey = None
                
            if assistant_2_jersey == 0:
                assistant_2_jersey = None
                

            # Append the data to data_to_save
            scores_to_s3.append({  
                "GameDate": game_date.strftime('%Y-%m-%d'),
                "Team": selected_team,
                "Opponent": opponent,
                "Home": stevenson_home, #Yes, No
                "Win": Game_Win, #Yes, No, Tie
                "ScoreStevenson":Score_Stevenson,
                "ScoreOpponent":Score_Opponent,
                "Period": period_game, 
                "Goal": score_jersey_number,               
                "Assistant_1": assistant_1_jersey,
                "Assistant_2": assistant_2_jersey
            })
            
    st.markdown("<hr>", unsafe_allow_html=True)
    
    # Add a "SAVE" button to save the data
    if st.button("Save Scores"):
        if not scores_to_s3:
            st.warning("No score data to save. Please add data before saving.")
        else:
                # Check if data_to_save is not empty, and write it to storage (or process it)
                if scores_to_s3:
                    try:
                        # Call the function to append data and upload back to S3
                        append_to_excel_s3(S3_BUCKET, EXCEL_FILE_KEY, scores_to_s3, "Scoring")


                        # Assuming df is currently a list
                        if isinstance(scores_to_s3, list):
                            # Case 1: List of lists
                            if all(isinstance(i, list) for i in scores_to_s3):
                                # Convert list of lists into DataFrame
                                scores_to_s3 = pd.DataFrame(scores_to_s3, columns=['GameDate', 'Team', 'Opponent','Period','JerseyNumber','Assistant_1','Assistant_2'])  # Adjust column names as needed

                            # Case 2: List of dictionaries
                            elif all(isinstance(i, dict) for i in scores_to_s3):
                                # Convert list of dictionaries into DataFrame
                                scores_to_s3 = pd.DataFrame(scores_to_s3)


                        #write backup file:
                        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                        file_name = f"temp/scores_{selected_team}_{current_time}.csv"

                        # Convert DataFrame to CSV in memory
                        csv_buffer = BytesIO()
                        scores_to_s3.to_csv(csv_buffer, index=False)

                        csv_buffer.seek(0)

                        # Upload the CSV file to S3
                        s3.put_object(Bucket=S3_BUCKET, Key=file_name, Body=csv_buffer.getvalue())



                        # Display success message
                        st.success("Score data successfully uploaded!")

                    except Exception as e:
                        # Display error message if something goes wrong
                        st.error(f"An error occurred while uploading the sscore data: {e}")                              
            
            
            

st.markdown("<hr>", unsafe_allow_html=True) 




# Expander for "Hockey Game Shots Input"
st.title("Hockey Game Shots Input")

with st.expander("Game Shots Input", expanded=False):
    #st.title("Hockey Game Shots Input")
            
    # Add a radio button for selecting the period, displayed horizontally
    period = st.radio("Select Period", options=["1", "2", "3", "Overtime"], horizontal=True)
    
    # Create two sections: Stevenson Team and Opponent Team, with a vertical line in between
    col1, col_mid, col2 = st.columns([10, 1, 10])  # Adjust column width ratios as needed
    
    # Stevenson Team Section
    with col1:
        st.header(f"Stevenson Team - Period {period}")
        stevenson_shots = st.slider(f"Number of Shots by Stevenson in Period {period}", min_value=0, max_value=30, value=0)
        
        if stevenson_shots > 0:
            st.subheader(f"Stevenson Shots Details for Period {period}")
            for i in range(stevenson_shots):
                cols = st.columns([2, 2])  # Define column widths
                with cols[0]:
                    shoot_zone = st.selectbox(f"Shoot Zone (Shot {i+1})", 
                                              ["1 - Inner Slot", "2 - West Outer Slot", "3 - East Outer Slot", "4 - Outside North West", "5 - Outside North East", "6 - West Point", "7 - Center Point","8 - East Point"], 
                                              key=f"stevenson_shoot_zone_{i}")
                with cols[1]:
                    jersey_number = st.selectbox(f"Jersey Number (Shot {i+1})", unique_jersey_numbers, key=f"stevenson_jersey_{i}")
                    
                data_to_save.append({
                    "GameDate": game_date.strftime('%Y-%m-%d'),
                    "Team": selected_team,
                    "Opponent": opponent,
                    "Period": period,
                    "JerseyNumber": jersey_number,               
                    "ShootingTeam": "Stevenson",
                    "ShootZone": shoot_zone
                })
    
    # Vertical line separator
    with col_mid:
        st.markdown("##")
        st.markdown(".")
                
    # Opponent Team Section
    with col2:
        st.header(f"Opponent Team - Period {period}")
        opponent_shots = st.slider(f"Number of Shots by Opponent in Period {period}", min_value=0, max_value=30, value=0)
        
        if opponent_shots > 0:
            st.subheader(f"Opponent Shots Details for Period {period}")
            for i in range(opponent_shots):
                cols = st.columns([2, 2])  # Define column widths
                with cols[0]:
                    shoot_zone = st.selectbox(f"Shoot Zone (Shot {i+1})", 
                                              ["1 - Inner Slot", "2 - West Outer Slot", "3 - East Outer Slot", "4 - Outside North West", "5 - Outside North East", "6 - West Point", "7 - Center Point","8 - East Point"], 
                                              key=f"opponent_shoot_zone_{i}")
                with cols[1]:
                    jersey_number = st.text_input(f"Jersey Number (Shot {i+1})", value="0", key=f"opponent_jersey_{i}")
                    
                data_to_save.append({
                    "GameDate": game_date.strftime('%Y-%m-%d'),
                    "Team": selected_team,
                    "Opponent": opponent,
                    "Period": period,        
                    "JerseyNumber": jersey_number,                
                    "ShootingTeam": opponent,
                    "ShootZone": shoot_zone
                })
                
    st.markdown("<hr>", unsafe_allow_html=True)                
                
    # Add a "SAVE" button to save the data
    if st.button("Save Shots"):
        if not data_to_save:
            st.warning("No data to save. Please add data before saving.")
        else:
                # Check if data_to_save is not empty, and write it to storage (or process it)
                if data_to_save:
                    try:
                        # Call the function to append data and upload back to S3
                        append_to_excel_s3(S3_BUCKET, EXCEL_FILE_KEY, data_to_save, "Shots")


                        # Assuming df is currently a list
                        if isinstance(data_to_save, list):
                            # Case 1: List of lists
                            if all(isinstance(i, list) for i in data_to_save):
                                # Convert list of lists into DataFrame
                                data_to_save = pd.DataFrame(data_to_save, columns=['GameDate', 'Team', 'Opponent','Period','JerseyNumber','ShootingTeam','ShootZone'])  # Adjust column names as needed

                            # Case 2: List of dictionaries
                            elif all(isinstance(i, dict) for i in data_to_save):
                                # Convert list of dictionaries into DataFrame
                                data_to_save = pd.DataFrame(data_to_save)


                        #write backup file:
                        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                        file_name = f"temp/shots_{selected_team}_{current_time}.csv"

                        # Convert DataFrame to CSV in memory
                        csv_buffer = BytesIO()
                        data_to_save.to_csv(csv_buffer, index=False)

                        csv_buffer.seek(0)

                        # Upload the CSV file to S3
                        s3.put_object(Bucket=S3_BUCKET, Key=file_name, Body=csv_buffer.getvalue())



                        # Display success message
                        st.success("Shots data successfully uploaded!")

                    except Exception as e:
                        # Display error message if something goes wrong
                        st.error(f"An error occurred while uploading the shots data: {e}")                



st.markdown("<hr>", unsafe_allow_html=True)  






# Expander for "Hockey Game Faceoff Input"
st.title("Hockey Game Faceoff Input")

with st.expander("Game Faceoff Input", expanded=False):
    #st.title("Hockey Game Faceoff Input")
    
    # Add a radio button for selecting the period, displayed horizontally
    period_faceoff = st.radio("Select Period", options=["1", "2", "3", "Overtime"], horizontal=True, key="period_faceoff")
    stevenson_faceoff = st.slider(f"Number of faceoff by Stevenson in Period {period_faceoff}", min_value=0, max_value=30, value=0)
    
    if stevenson_faceoff > 0:
        st.subheader(f"Stevenson Faceoff Details for Period {period_faceoff}")
        for i in range(stevenson_faceoff):
            # Define 3 columns, with equal or customizable widths
            cols = st.columns(3)  # Now, cols[0], cols[1], and cols[2] are valid
            
            with cols[0]:
                jersey_number = st.selectbox(f"Jersey Number (faceoff {i+1})", unique_jersey_numbers, key=f"faceoff_jersey_{i}")
            
            with cols[1]:
                user_input = st.text_input("Enter win count", key=f"numeric_input_{i}")

                # Initialize win_count as None
                win_count = None

                # Validate if the input is a number
                if user_input:
                    if user_input.isdigit():
                        win_count = int(user_input)  # Convert the input to an integer and store in win_count
                    else:
                        st.error("Please enter only numbers.")
            
            with cols[2]:
                lose_input = st.text_input("Enter lose count", key=f"numeric_lose_input_{i}")

                # Initialize lose_count as None
                lose_count = None

                # Validate if the input is a number
                if lose_input:
                    if lose_input.isdigit():
                        lose_count = int(lose_input)  # Convert the input to an integer and store in lose_count
                    else:
                        st.error("Please enter only numbers.")

            # Append the data to data_to_save
            faceoff_to_s3.append({  
                "GameDate": game_date.strftime('%Y-%m-%d'),
                "Team": selected_team,
                "Opponent": opponent,
                "Period": period_faceoff, 
                "JerseyNumber": jersey_number,               
                "Win": win_count,
                "Lose": lose_count
            })


    st.markdown("<hr>", unsafe_allow_html=True)            
            
            
    # Add a "SAVE" button to save the data
    if st.button("Save Faceoff"):
        if not faceoff_to_s3:
            st.warning("No data to save. Please add data before saving.")
        else:

                # Check if faceoff_to_s3 is not empty, and write it to storage (or process it)
                if faceoff_to_s3:
                    try:
                        # Call the function to append data and upload back to S3
                        append_to_excel_s3(S3_BUCKET, EXCEL_FILE_KEY, faceoff_to_s3, "Faceoff")


                        # Assuming df is currently a list
                        if isinstance(faceoff_to_s3, list):
                            # Case 1: List of lists
                            if all(isinstance(i, list) for i in faceoff_to_s3):
                                # Convert list of lists into DataFrame
                                faceoff_to_s3 = pd.DataFrame(faceoff_to_s3, columns=['GameDate', 'Team', 'Opponent','Period','JerseyNumber','Win','Lose'])  # Adjust column names as needed

                            # Case 2: List of dictionaries
                            elif all(isinstance(i, dict) for i in faceoff_to_s3):
                                # Convert list of dictionaries into DataFrame
                                faceoff_to_s3 = pd.DataFrame(faceoff_to_s3)


                        #write backup file:
                        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                        file_name = f"temp/faceoff_{selected_team}_{current_time}.csv"

                        # Convert DataFrame to CSV in memory
                        csv_buffer = BytesIO()
                        faceoff_to_s3.to_csv(csv_buffer, index=False)

                        csv_buffer.seek(0)

                        # Upload the CSV file to S3
                        s3.put_object(Bucket=S3_BUCKET, Key=file_name, Body=csv_buffer.getvalue())



                        # Display success message
                        st.success("Faceoff data successfully uploaded!")

                    except Exception as e:
                        # Display error message if something goes wrong
                        st.error(f"An error occurred while uploading the faceoff data: {e}")                
 

    
st.markdown("<hr>", unsafe_allow_html=True)      
    
    
    
# Expander for "Hockey Game Faceoff Input"
st.title("Hockey Game Goalie Data Input")

with st.expander("Goalie Data Input", expanded=False):
    #st.title("Hockey Game Faceoff Input")
    
    #goalie_jersey_numbers = unique_jersey_numbers
    
    #st.dataframe(unique_jersey_numbers, width=650)
    #st.dataframe(df_selected_team, width=650)    
    
    goalie_jersey_numbers = df_selected_team[df_selected_team['Position'] == 'Goalie']['JerseyNumber'].tolist()    
    
    
    cols = st.columns(3)  # Now, cols[0], cols[1], and cols[2] are valid
            
    with cols[0]:
        jersey_number = st.selectbox(f"Goalie Jersey Number", goalie_jersey_numbers, key=f"goalie_jersey")
            
    with cols[1]:
        user_input = st.text_input(f"{opponent} scores", key=f"opponentc_scores")

        # Initialize win_count as None
        opponentc_scores = 0

        # Validate if the input is a number
        if user_input:
            if user_input.isdigit():
                opponentc_scores = int(user_input)  # Convert the input to an integer and store in win_count
            else:
                st.error("Please enter only numbers.")
            
    with cols[2]:
        user_input = st.text_input(f"{opponent} shots", key=f"opponentc_shots")

        # Initialize win_count as None
        opponentc_shots = None

        # Validate if the input is a number
        if user_input:
            if user_input.isdigit():
                opponentc_shots = int(user_input)  # Convert the input to an integer and store in win_count
            else:
                st.error("Please enter only numbers.")

    # Append the data to data_to_save
    goalie_to_s3.append({  
                "GameDate": game_date.strftime('%Y-%m-%d'),
                "Team": selected_team,
                "Opponent": opponent,
                "JerseyNumber": jersey_number,               
                "Opponent_Score": opponentc_scores,
                "Opponent_Shots": opponentc_shots
        })

    st.markdown("<hr>", unsafe_allow_html=True)
    
    # Add a "SAVE" button to save the data
    if st.button("Save Goalie"):
        if not goalie_to_s3:
            st.warning("No data to save. Please add data before saving.")
        else:

                # Check if faceoff_to_s3 is not empty, and write it to storage (or process it)
                if goalie_to_s3:
                    try:
                        # Call the function to append data and upload back to S3
                        append_to_excel_s3(S3_BUCKET, EXCEL_FILE_KEY, goalie_to_s3, "Golie")


                        # Assuming df is currently a list
                        if isinstance(goalie_to_s3, list):
                            # Case 1: List of lists
                            if all(isinstance(i, list) for i in goalie_to_s3):
                                # Convert list of lists into DataFrame
                                goalie_to_s3 = pd.DataFrame(goalie_to_s3, columns=['GameDate', 'Team', 'Opponent','JerseyNumber','Opponent_Score','Opponent_Shots'])  # Adjust column names as needed

                            # Case 2: List of dictionaries
                            elif all(isinstance(i, dict) for i in goalie_to_s3):
                                # Convert list of dictionaries into DataFrame
                                goalie_to_s3 = pd.DataFrame(goalie_to_s3)


                        #write backup file:
                        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                        file_name = f"temp/goalie_{selected_team}_{current_time}.csv"

                        # Convert DataFrame to CSV in memory
                        csv_buffer = BytesIO()
                        goalie_to_s3.to_csv(csv_buffer, index=False)

                        csv_buffer.seek(0)

                        # Upload the CSV file to S3
                        s3.put_object(Bucket=S3_BUCKET, Key=file_name, Body=csv_buffer.getvalue())



                        # Display success message
                        st.success("Goalie data successfully uploaded!")

                    except Exception as e:
                        # Display error message if something goes wrong
                        st.error(f"An error occurred while uploading the goalie data: {e}")                
 
